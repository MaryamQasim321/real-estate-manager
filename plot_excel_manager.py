import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

DATA_DIR = "data/plots"


class PlotExcelManager:
    """
    Handles per-block Excel workbooks for the Plot Manager.
    Each block gets its own file: data/plots/<BlockName>.xlsx
      - "Plots" sheet:    PlotID | PlotNo | BuyerName | RatePerMarla | Marla | TotalPrice
      - "Payments" sheet: PaymentID | PlotID | Date | Description | AmountPaid
    Paid / Remaining are always derived from the Payments sheet, never stored,
    so the numbers can never drift out of sync.
    """

    PLOT_HEADERS = ["PlotID", "PlotNo", "BuyerName", "RatePerMarla", "Marla", "TotalPrice"]
    PAYMENT_HEADERS = ["PaymentID", "PlotID", "Date", "Description", "AmountPaid"]

    def __init__(self, data_dir=DATA_DIR):
        self.data_dir = data_dir
        os.makedirs(self.data_dir, exist_ok=True)

    # ------------------------------------------------------------------
    @staticmethod
    def _safe_name(block_name):
        keep = "-_() "
        cleaned = "".join(c for c in block_name.strip() if c.isalnum() or c in keep)
        return cleaned or "Block"

    def file_path(self, block_name):
        return os.path.join(self.data_dir, f"{self._safe_name(block_name)}.xlsx")

    # ------------------------------------------------------------------
    def _load_workbook(self, block_name):
        """Loads a block's workbook, migrating older files (created before the
        BuyerName column existed) so they keep working without losing data."""
        path = self.file_path(block_name)
        wb = load_workbook(path)
        sheet = wb["Plots"]
        headers = [cell.value for cell in sheet[1]]
        if "BuyerName" not in headers:
            sheet.insert_cols(3)
            sheet.cell(row=1, column=3, value="BuyerName")
            for r in range(2, sheet.max_row + 1):
                if sheet.cell(row=r, column=1).value is not None:
                    sheet.cell(row=r, column=3, value="")
            wb.save(path)
        return wb

    # ------------------------------------------------------------------
    # Block operations
    # ------------------------------------------------------------------
    def block_exists(self, block_name):
        return os.path.exists(self.file_path(block_name))

    def create_block(self, block_name):
        path = self.file_path(block_name)
        if os.path.exists(path):
            return False
        wb = Workbook()
        plots = wb.active
        plots.title = "Plots"
        plots.append(self.PLOT_HEADERS)
        payments = wb.create_sheet("Payments")
        payments.append(self.PAYMENT_HEADERS)
        wb.save(path)
        return True

    def delete_block(self, block_name):
        path = self.file_path(block_name)
        if os.path.exists(path):
            os.remove(path)

    def list_blocks(self):
        if not os.path.isdir(self.data_dir):
            return []
        return sorted(f[:-5] for f in os.listdir(self.data_dir) if f.endswith(".xlsx"))

    # ------------------------------------------------------------------
    # Plot operations
    # ------------------------------------------------------------------
    def add_plot(self, block_name, plot_no, buyer_name, rate_per_marla, marla):
        path = self.file_path(block_name)
        wb = self._load_workbook(block_name)
        sheet = wb["Plots"]

        existing_ids = [r[0].value for r in sheet.iter_rows(min_row=2) if r[0].value is not None]
        new_id = (max(existing_ids) + 1) if existing_ids else 1
        total_price = round(rate_per_marla * marla, 2)

        sheet.append([new_id, plot_no, buyer_name, rate_per_marla, marla, total_price])
        wb.save(path)
        return new_id

    def update_buyer_name(self, block_name, plot_id, buyer_name):
        path = self.file_path(block_name)
        wb = self._load_workbook(block_name)
        sheet = wb["Plots"]
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == plot_id:
                row[2].value = buyer_name
                break
        wb.save(path)

    def delete_plot(self, block_name, plot_id):
        path = self.file_path(block_name)
        wb = self._load_workbook(block_name)

        plots_sheet = wb["Plots"]
        for row in list(plots_sheet.iter_rows(min_row=2)):
            if row[0].value == plot_id:
                plots_sheet.delete_rows(row[0].row, 1)
                break

        payments_sheet = wb["Payments"]
        for row in list(payments_sheet.iter_rows(min_row=2)):
            if row[1].value == plot_id:
                payments_sheet.delete_rows(row[0].row, 1)

        wb.save(path)

    def load_plots(self, block_name):
        wb = self._load_workbook(block_name)
        plots_sheet = wb["Plots"]
        payments_sheet = wb["Payments"]

        paid_by_plot = {}
        for row in payments_sheet.iter_rows(min_row=2, values_only=True):
            if row[1] is None:
                continue
            paid_by_plot[row[1]] = paid_by_plot.get(row[1], 0) + (row[4] or 0)

        plots = []
        for row in plots_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            plot_id, plot_no, buyer_name, rate, marla, total_price = row
            paid = round(paid_by_plot.get(plot_id, 0), 2)
            remaining = round(total_price - paid, 2)
            plots.append({
                "id": plot_id, "plot_no": plot_no, "buyer_name": buyer_name or "",
                "rate": rate, "marla": marla,
                "total_price": total_price, "paid": paid, "remaining": remaining,
            })
        return plots

    # ------------------------------------------------------------------
    # Payment operations
    # ------------------------------------------------------------------
    def add_payment(self, block_name, plot_id, date_str, description, amount):
        path = self.file_path(block_name)
        wb = self._load_workbook(block_name)
        sheet = wb["Payments"]

        existing_ids = [r[0].value for r in sheet.iter_rows(min_row=2) if r[0].value is not None]
        new_id = (max(existing_ids) + 1) if existing_ids else 1

        sheet.append([new_id, plot_id, date_str, description, round(amount, 2)])
        wb.save(path)
        return new_id

    def delete_payment(self, block_name, payment_id):
        path = self.file_path(block_name)
        wb = self._load_workbook(block_name)
        sheet = wb["Payments"]
        for row in list(sheet.iter_rows(min_row=2)):
            if row[0].value == payment_id:
                sheet.delete_rows(row[0].row, 1)
                break
        wb.save(path)

    def load_payments(self, block_name, plot_id):
        wb = self._load_workbook(block_name)
        sheet = wb["Payments"]

        payments = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None or row[1] != plot_id:
                continue
            payments.append({
                "id": row[0], "plot_id": row[1], "date": row[2],
                "description": row[3], "amount": row[4] or 0,
            })
        return payments

    # ------------------------------------------------------------------
    def load_all_payments_for_month(self, year, month):
        """Returns every installment payment, across all blocks, made in a given month."""
        results = []
        for block in self.list_blocks():
            wb = self._load_workbook(block)

            plots_sheet = wb["Plots"]
            plot_no_by_id = {}
            buyer_by_id = {}
            for row in plots_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                plot_no_by_id[row[0]] = row[1]
                buyer_by_id[row[0]] = row[2] or ""

            payments_sheet = wb["Payments"]
            for row in payments_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is None or not row[2]:
                    continue
                payment_id, plot_id, date_str, description, amount = row
                try:
                    parsed = datetime.strptime(str(date_str), "%Y-%m-%d")
                except ValueError:
                    continue
                if parsed.year == year and parsed.month == month:
                    results.append({
                        "block": block,
                        "plot_no": plot_no_by_id.get(plot_id, "?"),
                        "buyer_name": buyer_by_id.get(plot_id, ""),
                        "date": date_str,
                        "description": description,
                        "amount": amount or 0,
                    })
        return results

    # ------------------------------------------------------------------
    def get_block_summary(self, block_name):
        plots = self.load_plots(block_name)
        total_price = round(sum(p["total_price"] for p in plots), 2)
        total_paid = round(sum(p["paid"] for p in plots), 2)
        total_remaining = round(total_price - total_paid, 2)
        return len(plots), total_price, total_paid, total_remaining

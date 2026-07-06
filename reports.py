import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import calendar
import os
import subprocess
import platform

from openpyxl import Workbook

from excel_manager import ExcelManager
from plot_excel_manager import PlotExcelManager

NAVY = "#1E3A5F"
BG = "#F5F7FA"
GREEN = "#28A745"
RED = "#DC3545"
BLUE = "#0D6EFD"
PURPLE = "#6F42C1"
BORDER = "#DDDDDD"

MONTH_NAMES = list(calendar.month_name)[1:]  # January .. December
REPORTS_DIR = "data/reports"


class Reports(ctk.CTkFrame):

    def __init__(self, parent):
        super().__init__(parent, fg_color=BG)
        self.pack(fill="both", expand=True)

        self.expense_excel = ExcelManager()
        self.plot_excel = PlotExcelManager()

        now = datetime.now()
        self.selected_year = now.year
        self.selected_month = now.month

        self._last_transactions = []
        self._last_installments = []
        self._last_totals = (0, 0, 0, 0)

        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=1)
        self.grid_columnconfigure(3, weight=1)
        self.grid_rowconfigure(2, weight=1)

        self.create_header()
        self.create_summary_cards()
        self.create_tables()

        self.refresh()

    # ------------------------------------------------------------------
    def create_header(self):
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.grid(row=0, column=0, columnspan=4, sticky="ew", padx=30, pady=(25, 15))
        header.grid_columnconfigure(0, weight=1)

        title_box = ctk.CTkFrame(header, fg_color="transparent")
        title_box.grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(title_box, text="Reports", font=("Arial", 32, "bold"), text_color=NAVY).pack(anchor="w")
        ctk.CTkLabel(title_box, text="Monthly summary of revenue, expenses, and installments",
                     font=("Arial", 16), text_color="gray40").pack(anchor="w")

        control_box = ctk.CTkFrame(header, fg_color="transparent")
        control_box.grid(row=0, column=1, sticky="e")

        self.month_var = tk.StringVar(value=MONTH_NAMES[self.selected_month - 1])
        self.year_var = tk.StringVar(value=str(self.selected_year))

        ctk.CTkOptionMenu(control_box, values=MONTH_NAMES, variable=self.month_var,
                           width=140, command=self.on_period_change).pack(side="left", padx=(0, 10))

        years = [str(y) for y in range(self.selected_year - 5, self.selected_year + 2)]
        ctk.CTkOptionMenu(control_box, values=years, variable=self.year_var,
                           width=100, command=self.on_period_change).pack(side="left", padx=(0, 10))

        ctk.CTkButton(control_box, text="Export to Excel", width=140, command=self.export_report).pack(side="left")

    def on_period_change(self, _=None):
        self.selected_month = MONTH_NAMES.index(self.month_var.get()) + 1
        self.selected_year = int(self.year_var.get())
        self.refresh()

    # ------------------------------------------------------------------
    def create_summary_cards(self):
        self.revenue_var = tk.StringVar(value="PKR 0")
        self.expense_var = tk.StringVar(value="PKR 0")
        self.balance_var = tk.StringVar(value="PKR 0")
        self.installments_var = tk.StringVar(value="PKR 0")

        self.make_card("Total Revenue", self.revenue_var, GREEN).grid(
            row=1, column=0, padx=(30, 10), pady=(0, 15), sticky="nsew")
        self.make_card("Total Expense", self.expense_var, RED).grid(
            row=1, column=1, padx=10, pady=(0, 15), sticky="nsew")
        self.make_card("Net Balance", self.balance_var, BLUE).grid(
            row=1, column=2, padx=10, pady=(0, 15), sticky="nsew")
        self.make_card("Installments Received", self.installments_var, PURPLE).grid(
            row=1, column=3, padx=(10, 30), pady=(0, 15), sticky="nsew")

    def make_card(self, title, value_var, color):
        card = ctk.CTkFrame(self, corner_radius=18, fg_color="white", border_width=1,
                             border_color=BORDER, height=110)
        ctk.CTkLabel(card, text=title, font=("Arial", 14, "bold"), text_color="gray35").pack(
            anchor="w", padx=18, pady=(16, 4))
        ctk.CTkLabel(card, textvariable=value_var, font=("Arial", 22, "bold"), text_color=color).pack(
            anchor="w", padx=18, pady=(0, 16))
        return card

    # ------------------------------------------------------------------
    def create_tables(self):
        container = ctk.CTkFrame(self, corner_radius=18, fg_color="white", border_width=1, border_color=BORDER)
        container.grid(row=2, column=0, columnspan=4, padx=30, pady=(0, 25), sticky="nsew")
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.tabview = ctk.CTkTabview(container, corner_radius=14,
                                       segmented_button_selected_color=NAVY,
                                       segmented_button_selected_hover_color=NAVY)
        self.tabview.grid(row=0, column=0, sticky="nsew", padx=15, pady=15)

        self.tab_revenue = self.tabview.add("Revenue & Expenses")
        self.tab_installments = self.tabview.add("Installments")

        self.tab_revenue.grid_rowconfigure(0, weight=1)
        self.tab_revenue.grid_columnconfigure(0, weight=1)
        self.tab_installments.grid_rowconfigure(0, weight=1)
        self.tab_installments.grid_columnconfigure(0, weight=1)

        style = ttk.Style()
        style.theme_use("default")
        style.configure("Report.Treeview", background="white", fieldbackground="white",
                         foreground="#333333", rowheight=30, borderwidth=0, font=("Arial", 12))
        style.configure("Report.Treeview.Heading", background="#ECECEC", foreground=NAVY,
                         font=("Arial", 12, "bold"), relief="flat")
        style.map("Report.Treeview", background=[("selected", "#CDE3FF")], foreground=[("selected", NAVY)])

        # ---- Revenue & Expense table ----
        rev_columns = ("date", "description", "type", "revenue", "expense", "balance")
        self.revenue_tree = ttk.Treeview(self.tab_revenue, columns=rev_columns, show="headings",
                                          style="Report.Treeview")
        rev_headers = {
            "date": ("Date", 100), "description": ("Description", 260), "type": ("Type", 90),
            "revenue": ("Revenue", 120), "expense": ("Expense", 120), "balance": ("Balance", 120),
        }
        for col, (text, width) in rev_headers.items():
            self.revenue_tree.heading(col, text=text)
            self.revenue_tree.column(col, width=width,
                                      anchor="w" if col in ("date", "description", "type") else "e")

        rev_scroll = ttk.Scrollbar(self.tab_revenue, orient="vertical", command=self.revenue_tree.yview)
        self.revenue_tree.configure(yscrollcommand=rev_scroll.set)
        self.revenue_tree.grid(row=0, column=0, sticky="nsew", pady=(10, 10))
        rev_scroll.grid(row=0, column=1, sticky="ns", pady=(10, 10))

        self.revenue_empty = ctk.CTkLabel(self.tab_revenue, text="No revenue or expense transactions this month.",
                                           font=("Arial", 13), text_color="gray50")

        # ---- Installments table ----
        inst_columns = ("block", "plot_no", "buyer", "date", "description", "amount")
        self.installment_tree = ttk.Treeview(self.tab_installments, columns=inst_columns, show="headings",
                                              style="Report.Treeview")
        inst_headers = {
            "block": ("Block", 110), "plot_no": ("Plot No.", 90), "buyer": ("Buyer Name", 150),
            "date": ("Date", 100), "description": ("Description", 220), "amount": ("Amount", 130),
        }
        for col, (text, width) in inst_headers.items():
            self.installment_tree.heading(col, text=text)
            self.installment_tree.column(col, width=width, anchor="w" if col != "amount" else "e")

        inst_scroll = ttk.Scrollbar(self.tab_installments, orient="vertical", command=self.installment_tree.yview)
        self.installment_tree.configure(yscrollcommand=inst_scroll.set)
        self.installment_tree.grid(row=0, column=0, sticky="nsew", pady=(10, 10))
        inst_scroll.grid(row=0, column=1, sticky="ns", pady=(10, 10))

        self.installment_empty = ctk.CTkLabel(self.tab_installments, text="No installments received this month.",
                                               font=("Arial", 13), text_color="gray50")

    # ------------------------------------------------------------------
    def refresh(self):
        # ---- Revenue & Expense data ----
        for row in self.revenue_tree.get_children():
            self.revenue_tree.delete(row)

        transactions = self.expense_excel.load_transactions(self.selected_year, self.selected_month)

        if not transactions:
            self.revenue_empty.grid(row=0, column=0, sticky="nsew", pady=(10, 10))
        else:
            self.revenue_empty.grid_forget()

        for t in transactions:
            self.revenue_tree.insert("", "end", values=(
                t["date"], t["description"], t["type"],
                f"PKR {t['revenue']:,.2f}" if t["revenue"] else "-",
                f"PKR {t['expense']:,.2f}" if t["expense"] else "-",
                f"PKR {t['balance']:,.2f}",
            ))

        total_revenue, total_expense, balance = self.expense_excel.get_summary(
            self.selected_year, self.selected_month)

        # ---- Installments data ----
        for row in self.installment_tree.get_children():
            self.installment_tree.delete(row)

        installments = self.plot_excel.load_all_payments_for_month(self.selected_year, self.selected_month)

        if not installments:
            self.installment_empty.grid(row=0, column=0, sticky="nsew", pady=(10, 10))
        else:
            self.installment_empty.grid_forget()

        for p in installments:
            self.installment_tree.insert("", "end", values=(
                p["block"], p["plot_no"], p["buyer_name"], p["date"], p["description"], f"PKR {p['amount']:,.2f}",
            ))

        total_installments = round(sum(p["amount"] for p in installments), 2)

        self.revenue_var.set(f"PKR {total_revenue:,.2f}")
        self.expense_var.set(f"PKR {total_expense:,.2f}")
        self.balance_var.set(f"PKR {balance:,.2f}")
        self.installments_var.set(f"PKR {total_installments:,.2f}")

        self._last_transactions = transactions
        self._last_installments = installments
        self._last_totals = (total_revenue, total_expense, balance, total_installments)

    # ------------------------------------------------------------------
    def export_report(self):
        os.makedirs(REPORTS_DIR, exist_ok=True)

        month_name = MONTH_NAMES[self.selected_month - 1]
        filename = f"Report_{self.selected_year}-{self.selected_month:02d}.xlsx"
        path = os.path.join(REPORTS_DIR, filename)

        wb = Workbook()

        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        summary_sheet.append(["Property Management System - Monthly Report"])
        summary_sheet.append([f"{month_name} {self.selected_year}"])
        summary_sheet.append([])
        summary_sheet.append(["Metric", "Amount (PKR)"])

        total_revenue, total_expense, balance, total_installments = self._last_totals
        summary_sheet.append(["Total Revenue", total_revenue])
        summary_sheet.append(["Total Expense", total_expense])
        summary_sheet.append(["Net Balance", balance])
        summary_sheet.append(["Installments Received", total_installments])
        summary_sheet.append(["Grand Total Received", round(total_revenue + total_installments, 2)])

        summary_sheet.column_dimensions["A"].width = 32
        summary_sheet.column_dimensions["B"].width = 20

        rev_sheet = wb.create_sheet("Revenue & Expenses")
        rev_sheet.append(["Date", "Description", "Type", "Revenue", "Expense", "Balance"])
        for t in self._last_transactions:
            rev_sheet.append([t["date"], t["description"], t["type"], t["revenue"], t["expense"], t["balance"]])
        for col_letter, width in zip("ABCDEF", (14, 30, 12, 14, 14, 14)):
            rev_sheet.column_dimensions[col_letter].width = width

        inst_sheet = wb.create_sheet("Installments")
        inst_sheet.append(["Block", "Plot No.", "Buyer Name", "Date", "Description", "Amount"])
        for p in self._last_installments:
            inst_sheet.append([p["block"], p["plot_no"], p["buyer_name"], p["date"], p["description"], p["amount"]])
        for col_letter, width in zip("ABCDEF", (16, 14, 20, 14, 30, 14)):
            inst_sheet.column_dimensions[col_letter].width = width

        wb.save(path)

        if messagebox.askyesno(
                "Report Exported",
                f"Report saved to:\n{os.path.abspath(path)}\n\nOpen the containing folder now?"):
            self.open_folder(os.path.abspath(REPORTS_DIR))

    @staticmethod
    def open_folder(path):
        try:
            system = platform.system()
            if system == "Windows":
                os.startfile(path)
            elif system == "Darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])
        except Exception:
            pass

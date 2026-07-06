import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

DATA_DIR = "data"


class ExcelManager:
    """
    Handles all Excel read/write operations for the Expense Tracker.
    Every calendar month gets its own workbook: data/YYYY-MM.xlsx
    Sheet "Transactions" columns: ID | Date | Description | Type | Revenue | Expense | Balance
    """

    HEADERS = ["ID", "Date", "Description", "Type", "Revenue", "Expense", "Balance"]

    def __init__(self, data_dir=DATA_DIR):
        self.data_dir = data_dir
        os.makedirs(self.data_dir, exist_ok=True)

    # ------------------------------------------------------------------
    def file_path(self, year, month):
        return os.path.join(self.data_dir, f"{year}-{month:02d}.xlsx")

    def _ensure_file(self, year, month):
        path = self.file_path(year, month)
        if not os.path.exists(path):
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Transactions"
            sheet.append(self.HEADERS)
            for i, header in enumerate(self.HEADERS, start=1):
                sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = 16
            wb.save(path)
        return path

    # ------------------------------------------------------------------
    def load_transactions(self, year, month):
        path = self._ensure_file(year, month)
        wb = load_workbook(path)
        sheet = wb["Transactions"]

        transactions = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            transactions.append({
                "id": row[0],
                "date": row[1],
                "description": row[2],
                "type": row[3],
                "revenue": row[4] or 0,
                "expense": row[5] or 0,
                "balance": row[6] or 0,
            })
        return transactions

    # ------------------------------------------------------------------
    def add_transaction(self, year, month, date_str, description, ttype, amount):
        path = self._ensure_file(year, month)
        wb = load_workbook(path)
        sheet = wb["Transactions"]

        existing_ids = [r[0].value for r in sheet.iter_rows(min_row=2) if r[0].value is not None]
        new_id = (max(existing_ids) + 1) if existing_ids else 1

        prev_balance = sheet.cell(row=sheet.max_row, column=7).value if sheet.max_row > 1 else 0
        prev_balance = prev_balance or 0

        revenue = round(amount, 2) if ttype == "Revenue" else 0
        expense = round(amount, 2) if ttype == "Expense" else 0
        new_balance = round(prev_balance + revenue - expense, 2)

        sheet.append([new_id, date_str, description, ttype, revenue, expense, new_balance])
        wb.save(path)
        return new_balance

    # ------------------------------------------------------------------
    def delete_transaction(self, year, month, transaction_id):
        path = self._ensure_file(year, month)
        wb = load_workbook(path)
        sheet = wb["Transactions"]

        for row in list(sheet.iter_rows(min_row=2)):
            if row[0].value == transaction_id:
                sheet.delete_rows(row[0].row, 1)
                break

        # Recalculate the running balance top to bottom after deletion
        balance = 0
        for row in sheet.iter_rows(min_row=2):
            revenue = row[4].value or 0
            expense = row[5].value or 0
            balance = round(balance + revenue - expense, 2)
            row[6].value = balance

        wb.save(path)

    # ------------------------------------------------------------------
    def get_summary(self, year, month):
        transactions = self.load_transactions(year, month)
        total_revenue = round(sum(t["revenue"] for t in transactions), 2)
        total_expense = round(sum(t["expense"] for t in transactions), 2)
        balance = round(total_revenue - total_expense, 2)
        return total_revenue, total_expense, balance

    # ------------------------------------------------------------------
    def available_months(self):
        """Sorted list of (year, month) tuples that currently have a workbook."""
        months = []
        if os.path.isdir(self.data_dir):
            for fname in os.listdir(self.data_dir):
                if fname.endswith(".xlsx") and len(fname) == 12:
                    try:
                        year, month = fname[:-5].split("-")
                        months.append((int(year), int(month)))
                    except ValueError:
                        continue
        if not months:
            now = datetime.now()
            months.append((now.year, now.month))
        return sorted(set(months))
